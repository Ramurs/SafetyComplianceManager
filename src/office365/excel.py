from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.config import get_settings


def generate_audit_excel(audit) -> Path:
    """Generate an Excel spreadsheet for audit findings."""
    settings = get_settings()
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Findings"

    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = ["Severity", "Control ID", "Finding Title", "Description", "Recommendation", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    severity_fills = {
        "critical": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "high": PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"),
        "medium": PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid"),
        "low": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
        "info": PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid"),
    }

    if audit.findings:
        for row, finding in enumerate(audit.findings, 2):
            ws.cell(row=row, column=1, value=finding.severity.upper()).fill = severity_fills.get(finding.severity, PatternFill())
            ws.cell(row=row, column=2, value=finding.control_id)
            ws.cell(row=row, column=3, value=finding.title)
            ws.cell(row=row, column=4, value=finding.description)
            ws.cell(row=row, column=5, value=finding.recommendation)
            ws.cell(row=row, column=6, value=finding.status)

    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    output_path = settings.output_dir / f"audit_{audit.id[:8]}.xlsx"
    wb.save(str(output_path))
    return output_path


def generate_risk_register(risks: list) -> Path:
    """Generate an Excel risk register."""
    settings = get_settings()
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Register"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = ["Risk ID", "Title", "Category", "Likelihood", "Impact", "Score", "Status", "Owner", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row, risk in enumerate(risks, 2):
        ws.cell(row=row, column=1, value=risk.id[:8])
        ws.cell(row=row, column=2, value=risk.title)
        ws.cell(row=row, column=3, value=risk.category)
        ws.cell(row=row, column=4, value=risk.likelihood)
        ws.cell(row=row, column=5, value=risk.impact)

        score_cell = ws.cell(row=row, column=6, value=risk.score)
        if risk.score >= 16:
            score_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        elif risk.score >= 9:
            score_cell.fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        ws.cell(row=row, column=7, value=risk.status)
        ws.cell(row=row, column=8, value=risk.owner)
        ws.cell(row=row, column=9, value=risk.description)

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    output_path = settings.output_dir / "risk_register.xlsx"
    wb.save(str(output_path))
    return output_path
